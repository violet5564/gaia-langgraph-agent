from pathlib import Path
from typing import Annotated, TypedDict
from urllib.parse import urlparse
import os

from dotenv import load_dotenv
from langchain_core.messages import AnyMessage, HumanMessage, SystemMessage
from langchain_google_genai import ChatGoogleGenerativeAI
from langgraph.graph import START, StateGraph
from langgraph.graph.message import add_messages
from langgraph.prebuilt import ToolNode, tools_condition


load_dotenv()

GEMINI_MODEL = "gemini-3-flash-preview"
llm = ChatGoogleGenerativeAI(model=GEMINI_MODEL, temperature = 0)

GAIA_API_URL = "https://agents-course-unit4-scoring.hf.space"
DOWNLOAD_DIR = Path("gaia_files")


def safe_file_name(file_name: str) -> str:
    """Make a file name safe enough to write inside the local download folder."""
    cleaned_name = "".join(
        char if char.isalnum() or char in {".", "-", "_"} else "_"
        for char in file_name
    ).strip("._")
    return cleaned_name or "downloaded_file"


def calculator(expression: str) -> str:
    """Evaluate a simple math expression and return the result."""
    allowed_chars = set("0123456789+-*/(). %")
    if not expression or any(char not in allowed_chars for char in expression):
        return "Invalid expression. Only numbers and basic operators are allowed."

    try:
        return str(eval(expression, {"__builtins__": {}}, {}))
    except Exception as error:
        return f"Calculator error: {error}"


def read_text_file(file_path: str) -> str:
    """Read a local text file and return its content."""
    path = Path(file_path)
    if not path.exists():
        return f"File not found: {file_path}"
    if not path.is_file():
        return f"Not a file: {file_path}"

    try:
        return path.read_text(encoding="utf-8")[:8000]
    except UnicodeDecodeError:
        return path.read_text(encoding="gbk", errors="ignore")[:8000]
    except Exception as error:
        return f"File read error: {error}"


def require_existing_file(file_path: str) -> Path | str:
    """Return a Path for an existing file, or an error message string."""
    path = Path(file_path)
    if not path.exists():
        return f"File not found: {file_path}"
    if not path.is_file():
        return f"Not a file: {file_path}"
    return path


def download_gaia_file(task_id: str, file_name: str = "") -> str:
    """Download the attached GAIA file for a task and return the local file path."""
    if not task_id:
        return "Missing task_id."

    try:
        import requests
    except ImportError:
        return "Missing package. Run: pip install requests"

    DOWNLOAD_DIR.mkdir(exist_ok=True)

    try:
        response = requests.get(f"{GAIA_API_URL}/files/{task_id}", timeout=60)
        response.raise_for_status()
    except Exception as error:
        return f"GAIA file download error: {error}"

    if not file_name:
        content_disposition = response.headers.get("content-disposition", "")
        if "filename=" in content_disposition:
            file_name = content_disposition.split("filename=", 1)[1].strip('" ')
        else:
            file_name = f"{task_id}.bin"

    local_path = DOWNLOAD_DIR / safe_file_name(file_name)
    local_path.write_bytes(response.content)

    return f"Downloaded file to: {local_path}"


def read_pdf_file(file_path: str) -> str:
    """Extract text from a local PDF file."""
    path_or_error = require_existing_file(file_path)
    if isinstance(path_or_error, str):
        return path_or_error

    try:
        from pypdf import PdfReader
    except ImportError:
        return "Missing package. Run: pip install pypdf"

    try:
        reader = PdfReader(str(path_or_error))
        pages = []
        for page_number, page in enumerate(reader.pages, start=1):
            page_text = page.extract_text() or ""
            if page_text.strip():
                pages.append(f"--- Page {page_number} ---\n{page_text.strip()}")
        text = "\n\n".join(pages)
        return text[:16000] if text else "No extractable text found in PDF."
    except Exception as error:
        return f"PDF read error: {error}"


def read_excel_file(file_path: str) -> str:
    """Read a local .xlsx file and return visible sheet values as text."""
    path_or_error = require_existing_file(file_path)
    if isinstance(path_or_error, str):
        return path_or_error

    try:
        from openpyxl import load_workbook
    except ImportError:
        return "Missing package. Run: pip install openpyxl"

    try:
        workbook = load_workbook(path_or_error, data_only=True, read_only=True)
        sheet_outputs = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value) for value in row]
                if any(value.strip() for value in values):
                    rows.append("\t".join(values))
            if rows:
                sheet_outputs.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows[:200]))
        text = "\n\n".join(sheet_outputs)
        return text[:16000] if text else "No visible values found in Excel file."
    except Exception as error:
        return f"Excel read error: {error}"


def analyze_image_file(file_path: str, question: str = "Describe the image and extract any relevant text.") -> str:
    """Analyze a local image file using a multimodal Gemini model."""
    path_or_error = require_existing_file(file_path)
    if isinstance(path_or_error, str):
        return path_or_error

    try:
        from google import genai
    except ImportError:
        return "Missing package. Run: pip install google-genai"

    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        return "Missing GEMINI_API_KEY in environment."

    try:
        client = genai.Client(api_key=api_key)
        uploaded_file = client.files.upload(file=str(path_or_error))
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[question, uploaded_file],
        )
        return (response.text or "").strip()[:12000]
    except Exception as error:
        return f"Image analysis error: {error}"


def transcribe_audio_file(file_path: str, question: str = "Transcribe this audio accurately.") -> str:
    """Transcribe or analyze a local audio file using a multimodal Gemini model."""
    path_or_error = require_existing_file(file_path)
    if isinstance(path_or_error, str):
        return path_or_error

    try:
        from google import genai
    except ImportError:
        return "Missing package. Run: pip install google-genai"

    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        return "Missing GEMINI_API_KEY in environment."

    try:
        client = genai.Client(api_key=api_key)
        uploaded_file = client.files.upload(file=str(path_or_error))
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=[question, uploaded_file],
        )
        return (response.text or "").strip()[:12000]
    except Exception as error:
        return f"Audio transcription error: {error}"


def web_search(query: str) -> str:
    """Search the web for current or factual information."""
    try:
        from ddgs import DDGS
    except ImportError:
        return "ddgs is not installed. Run: pip install ddgs"

    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=5))
    except Exception as error:
        return f"Web search error: {error}"

    if not results:
        return "No search results found."

    formatted_results = []
    for index, result in enumerate(results, start=1):
        title = result.get("title", "No title")
        body = result.get("body", "No snippet")
        href = result.get("href", "No URL")
        formatted_results.append(f"{index}. {title}\n{body}\nURL: {href}")

    return "\n\n".join(formatted_results)


def fetch_webpage(url: str) -> str:
    """Fetch a webpage URL and return readable text from the page body."""
    parsed_url = urlparse(url)
    if parsed_url.scheme not in {"http", "https"}:
        return "Invalid URL. Only http and https URLs are supported."

    try:
        import requests
        from bs4 import BeautifulSoup
    except ImportError:
        return "Missing packages. Run: pip install requests beautifulsoup4"

    try:
        response = requests.get(
            url,
            headers={"User-Agent": "Mozilla/5.0 GAIA learning agent"},
            timeout=15,
        )
        response.raise_for_status()
    except Exception as error:
        return f"Webpage fetch error: {error}"

    soup = BeautifulSoup(response.text, "html.parser")

    # Remove page chrome that usually hurts answer extraction.
    for tag in soup(["script", "style", "nav", "footer", "header", "aside"]):
        tag.decompose()

    title = soup.title.get_text(" ", strip=True) if soup.title else "No title"
    text = soup.get_text("\n", strip=True)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    page_text = "\n".join(lines)

    # Keep tool output short enough for the model context.
    return f"Title: {title}\nURL: {url}\n\n{page_text[:12000]}"


tools = [
    calculator,
    read_text_file,
    download_gaia_file,
    read_pdf_file,
    read_excel_file,
    analyze_image_file,
    transcribe_audio_file,
    web_search,
    fetch_webpage,
]

llm_with_tools = llm.bind_tools(tools)


class AgentState(TypedDict):
    messages: Annotated[list[AnyMessage], add_messages]


def assistant(state: AgentState):
    sys_msg = SystemMessage(
        content="""
You are a GAIA-style research agent.

Core rules:
- Use tools when you need exact facts, current information, file contents, or arithmetic.
- Use web_search to find candidate URLs, then use fetch_webpage to read the most relevant source.
- If the question has an attached file, use download_gaia_file with the provided task_id before reading or analyzing it.
- For PDFs use read_pdf_file. For .xlsx files use read_excel_file. For images use analyze_image_file. For .mp3 audio use transcribe_audio_file.
- Do not call the same file tool repeatedly on the same file unless the previous tool result clearly failed.
- If transcribe_audio_file returns a transcript or useful audio content, immediately answer from that content instead of downloading or transcribing the same audio again.
- If an attached file has already been downloaded, reuse the local file path from the previous tool result.
- Do not guess factual answers.
- Today is 2026-04-28.
- For latest/current questions, search for information from 2026 first.
- If a required tool fails, say that the tool failed and do not provide an unverified answer.

Final answer rules:
- Provide only the final answer, not your reasoning process.
- Be concise and exact.
- Preserve required units, dates, names, capitalization, and spelling.
- If multiple items are requested, use a numbered list.
- If the question asks for a specific format, follow that format exactly.
- Do not include sources, URLs, explanations, markdown headings, or extra sentences unless the question explicitly asks for them.
- For comma separated list questions, output only the comma separated list.
- If you cannot verify the answer because a required tool failed, say so clearly.
"""
    )
    response = llm_with_tools.invoke([sys_msg] + state["messages"])
    return {"messages": [response]}


builder = StateGraph(AgentState)
builder.add_node("assistant", assistant)
builder.add_node("tools", ToolNode(tools))

builder.add_edge(START, "assistant")
builder.add_conditional_edges("assistant", tools_condition)
builder.add_edge("tools", "assistant")

react_graph = builder.compile()


if __name__ == "__main__":
    messages = [
        HumanMessage(
            # content="Calculate this exactly: first divide 6790 by 5, then multiply by 3, then add 100."
            # content=("read the file 'requrements.txt' and summarize its content in 2 sentences.")
            content = ("Search the web for 'latest news on AI advancements' and summarize the top 3 results in 2 sentences.")
        )
    ]
    try:
        result = react_graph.invoke(
            {"messages": messages},
            config={"recursion_limit": 8},
        )

        for message in result["messages"]:
            message.pretty_print()
    except Exception as error:
        print(f"Agent run failed: {error}")
